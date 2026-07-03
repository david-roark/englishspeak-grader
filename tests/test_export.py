"""Test xuất Excel."""

from __future__ import annotations

from openpyxl import load_workbook

from core.export import _safe_sheet_title, _slug, export_to_excel


def test_export_tạo_file_và_sheet(sample_result, tmp_path):
    out = export_to_excel(sample_result, video_name="bai1.mp4",
                          rubric_name="IELTS Speaking", out_path=tmp_path / "kq.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    # Sheet tổng quan + 1 sheet mỗi học sinh
    assert "Tổng quan" in wb.sheetnames
    assert len(wb.sheetnames) == 1 + len(sample_result.students)


def test_export_ghi_đúng_dữ_liệu_tổng_quan(sample_result, tmp_path):
    out = export_to_excel(sample_result, video_name="bai1.mp4",
                          rubric_name="IELTS Speaking", out_path=tmp_path / "kq.xlsx")
    ws = load_workbook(out)["Tổng quan"]
    assert ws["B3"].value == "bai1.mp4"
    assert ws["B4"].value == "IELTS Speaking"
    assert ws["B6"].value == "English"
    # Dòng học sinh đầu tiên (header ở dòng 9, dữ liệu từ dòng 10)
    assert ws["A10"].value == "An"
    assert ws["B10"].value == 6.5


def test_export_tự_tạo_tên_file_khi_không_truyền_path(sample_result):
    out = export_to_excel(sample_result, video_name="my video.mp4", rubric_name="R")
    try:
        assert out.exists()
        assert out.suffix == ".xlsx"
    finally:
        out.unlink(missing_ok=True)


def test_safe_sheet_title_giới_hạn_31_ký_tự():
    title = _safe_sheet_title("A" * 100, 1)
    assert len(title) <= 31
    assert title.startswith("1. ")


def test_safe_sheet_title_loại_ký_tự_cấm():
    title = _safe_sheet_title("Tên/có:ký*tự[cấm]", 2)
    assert not any(ch in title for ch in ':\\/?*[]')


def test_safe_sheet_title_tên_rỗng_dùng_nhãn():
    assert _safe_sheet_title("   ", 3) == "3. HS 3"


def test_slug():
    # Ký tự alnum (kể cả có dấu) được giữ; khoảng trắng/ký tự lạ -> "_"
    assert _slug("bai noi 1") == "bai_noi_1"
    assert _slug("a/b*c") == "a_b_c"
    assert _slug("") == "ket_qua"
    assert _slug("!!!") == "ket_qua"
