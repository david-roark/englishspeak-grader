"""Xuất kết quả chấm điểm ra file Excel (.xlsx)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.schemas import AssessmentResult

EXPORT_DIR = Path(__file__).resolve().parent.parent / "exports"

_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _bullets(items: list[str]) -> str:
    return "\n".join(f"• {x}" for x in items) if items else ""


def export_to_excel(
    result: AssessmentResult,
    *,
    video_name: str = "",
    rubric_name: str = "",
    out_path: str | Path | None = None,
) -> Path:
    """Tạo file Excel gồm: sheet Tổng quan + mỗi học sinh 1 sheet chi tiết.

    Trả về đường dẫn file đã ghi.
    """
    wb = Workbook()

    # -- Sheet tổng quan --------------------------------------------------- #
    ws = wb.active
    assert ws is not None  # Workbook mới luôn có sheet active
    ws.title = "Tổng quan"
    ws["A1"] = "KẾT QUẢ CHẤM BÀI NÓI TIẾNG ANH"
    ws["A1"].font = _TITLE_FONT
    ws["A3"] = "Video:"
    ws["B3"] = video_name
    ws["A4"] = "Rubric:"
    ws["B4"] = rubric_name
    ws["A5"] = "Ngày chấm:"
    ws["B5"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A6"] = "Ngôn ngữ nói:"
    ws["B6"] = result.detected_language
    ws["A7"] = "Tóm tắt:"
    ws["B7"] = result.task_summary
    ws["B7"].alignment = _WRAP
    for col in ("A",):
        for r in range(3, 8):
            ws[f"{col}{r}"].font = Font(bold=True)

    # Bảng tổng điểm các học sinh
    header_row = 9
    headers = ["Học sinh", "Tổng điểm", "Điểm tối đa", "Xếp loại", "Tên chắc chắn?"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    for i, s in enumerate(result.students, 1):
        row = header_row + i
        ws.cell(row=row, column=1, value=s.student_name)
        ws.cell(row=row, column=2, value=s.total_score)
        ws.cell(row=row, column=3, value=s.max_total_score)
        ws.cell(row=row, column=4, value=s.overall_level)
        ws.cell(row=row, column=5, value="Có" if s.name_confident else "Chưa (cần gán)")

    _autosize(ws, {"A": 22, "B": 40})

    # -- Mỗi học sinh 1 sheet chi tiết ------------------------------------- #
    for idx, s in enumerate(result.students, 1):
        title = _safe_sheet_title(s.student_name, idx)
        d = wb.create_sheet(title=title)
        d["A1"] = s.student_name
        d["A1"].font = _TITLE_FONT
        d["A2"] = f"Xếp loại: {s.overall_level}   |   Tổng: {s.total_score}/{s.max_total_score}"

        # Bảng điểm tiêu chí
        r = 4
        for j, h in enumerate(["Tiêu chí", "Điểm", "Tối đa", "Nhận xét"], 1):
            c = d.cell(row=r, column=j, value=h)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
        for cs in s.criteria:
            r += 1
            d.cell(row=r, column=1, value=cs.criterion)
            d.cell(row=r, column=2, value=cs.score)
            d.cell(row=r, column=3, value=cs.max_score)
            cell = d.cell(row=r, column=4, value=cs.comment)
            cell.alignment = _WRAP

        # Các mục nhận xét
        r += 2
        for label, items in [
            ("Điểm mạnh", s.strengths),
            ("Điểm yếu", s.weaknesses),
            ("Gợi ý cải thiện", s.improvement_suggestions),
            ("Hướng luyện tập", s.practice_directions),
        ]:
            d.cell(row=r, column=1, value=label).font = Font(bold=True)
            cell = d.cell(row=r, column=2, value=_bullets(items))
            cell.alignment = _WRAP
            r += 1

        # Dẫn chứng
        if s.evidence:
            r += 1
            d.cell(row=r, column=1, value="Dẫn chứng (timestamp)").font = Font(bold=True)
            r += 1
            for j, h in enumerate(["Thời điểm", "Trích dẫn", "Ghi chú"], 1):
                c = d.cell(row=r, column=j, value=h)
                c.fill = _HEADER_FILL
                c.font = _HEADER_FONT
            for ev in s.evidence:
                r += 1
                d.cell(row=r, column=1, value=ev.timestamp)
                d.cell(row=r, column=2, value=ev.quote).alignment = _WRAP
                d.cell(row=r, column=3, value=ev.note).alignment = _WRAP

        _autosize(d, {"A": 26, "B": 20, "C": 10, "D": 50})

    # -- Ghi file ---------------------------------------------------------- #
    if out_path is None:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = Path(video_name).stem if video_name else "ket_qua"
        out_path = EXPORT_DIR / f"{_slug(stem)}_{stamp}.xlsx"
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _safe_sheet_title(name: str, idx: int) -> str:
    # Excel: tên sheet <=31 ký tự, không chứa : \ / ? * [ ]
    bad = ':\\/?*[]'
    cleaned = "".join(ch for ch in name if ch not in bad).strip() or f"HS {idx}"
    title = cleaned[:28]
    return f"{idx}. {title}"[:31]


def _slug(text: str) -> str:
    keep = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in text)
    return keep.strip("_") or "ket_qua"
